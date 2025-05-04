import re
import shutil
import threading
import pandas as pd
from time import sleep
from math import ceil
from common_utils import *
from f5_cfg import *
from csm import *
from csm_cfg import *
from string import Template
from openpyxl import load_workbook
import wexpect
from concurrent.futures import ThreadPoolExecutor

"""
csm
    if mode in ['all']:
        csm_get_all_shared_policy(**kwargs)
        csm_get_all_policy_objects(**kwargs)

def csm_get_all_policy_objects(**kwargs):
    data_dir = kwargs.get('dir')
    for policy_obj_type in policy_obj_types:
        result = csm_GetPolicyObjectsListByType(obj_type=policy_obj_type)
        fname = data_dir + fr'\csm.{policy_obj_type}.xml'
        str2file(file=fname, str=result)

def csm_get_all_shared_policy(**kwargs):
    data_dir = kwargs.get('dir')
    policy_types = get_policy_types()
    for policy_type in policy_types:
        result = csm_GetSharedPolicyNamesByType(policy_type=policy_type)
        if 'baseError' not in result:
            fname = data_dir + fr'\csm.{policy_type}.xml'
            str2file(file=fname, str=result)


def csm_all_dev_policy_2_sheet(**kwargs):
    gid_dev_items = kwargs.get('gid_dev_items')
    dev_regex = kwargs.get('dev_regex')
    data_dir = kwargs.get('dir')
    for gid, dev in gid_dev_items:
        if re.search(dev_exclude_regex, dev) or (dev_regex and not re.search(dev_regex, dev)):
            continue
        csm_dev_policy_2_sheet(**kwargs, dev=dev, gid=gid)
    [f.unlink() for f in Path(data_dir).glob("*.xml")]

                csm_all_dev_policy_2_sheet(dir=data_dir, dev_regex=dev_regex, gid_dev_items=csm_get_gid_dev_items(dir=data_dir))

def csm_GetSharedPolicyNamesByType(**kwargs):
    url_func = 'getSharedPolicyListByType'
    xml_func = 'policyNamesByTypeRequest'
    policy_type = kwargs.get('policy_type')
    xml_args = f'<policyType>{policy_type}</policyType>'
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args)

def obj_type_name_2_xml_args(obj_type_names):
    xml_args = ''
    for obj_type, obj_name_list in obj_type_names.items():
        obj_name_str = ''
        for obj_name in obj_name_list:
            obj_name_str += f'<name>{obj_name}</name>'
        xml_args += f'<{obj_type}>{obj_name_str}</{obj_type}>'
    return xml_args

def csm_GetPolicyObject(**kwargs):
    url_func = 'getPolicyObject'
    xml_func = 'getPolicyObjectRequest'
    obj_type_names = kwargs.get('obj_type_names')
    xml_args = obj_type_name_2_xml_args(obj_type_names)
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args)
def csm_GetPolicyObjectsListByType(**kwargs):
    url_func = 'getPolicyObjectsListByType'
    xml_func = 'policyObjectsListByTypeRequest'
    obj_type = kwargs.get('obj_type')
    xml_args = f'<policyObjectType>{obj_type}</policyObjectType>'
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args)
                            
"""
def run_thread_parallel(task, *args, **kwargs):  # map args, kwargs const
    max_workers = kwargs.get('max_workers', 32)
    task_args = [(arg, kwargs) for arg in args]
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = list(executor.map(lambda p: task(p[0], **p[1]), task_args))  # iterator
    return results

# python_exe = fr'"C:\Program Files\Python313\python.exe"'
# for suffix in ['hit.vs.csm', 'acl.hit.2.rule', 'acl.hit']:
# acl_start_regex = r"access-list (?P<name>.+); \d+ elements; name hash: .+"
"""
        my_logger.info(f"{dev} no policy")
        my_logger.info(f"{dev} no {xml_file.name}")
        my_logger.info(f"{dev} no policy")
        my_logger.info(f"{dev} no policy")

                my_logger.info(f"{dev} no {policy_type}")
    debug = kwargs.get('debug', False)
    if debug:
        print(_url)
        print(_xml)
                print(f"\n{url_func} successful!")
                    print(f"cookie: {cookie}")
            print(f'\nSave cookie to file: {cookie}')
            print(f'\nLoad cookie from file: {cookie}')
            print('\nNo cookie')
    print(f"add cookie to session: {cookie}")
print(f"\n{dev} policy types:\n{dev_policy_types}")
            print(f"{Path(fname).name} up to date")
print(f'{Path(xml_file).name} -> {Path(save_file).name}')
                print(f"{Path(f).name} up to date")
    print(f"\n{dev} policy types:\n{dev_policy_types}")
        print(f'{Path(excel_file).name}')
                print(f'{sheet_name}')
            print(f'skip {dev}')
            print(f"{Path(excel_file).name} up to date")
        print(f"{Path(fname).name} up to date")

"""
def copy_files(**kwargs):
    dst_dir = kwargs.get('dst_dir')
    src_files = kwargs.get('src_files')
    for file in glob(src_files):
        shutil.copy(file, dst_dir)

def check_tab_in_excel(**kwargs):
    file = kwargs.get('file')
    tab = kwargs.get('tab')    
    return set(tab if isinstance(tab, list) else [tab]).issubset(load_workbook(file).sheetnames)        

def get_tab_pos(**kwargs):
    excel_file = kwargs.get('file')
    tab = kwargs.get('tab')
    book = load_workbook(excel_file)
    return book._sheets.index(book[tab])

def move_tab(**kwargs):
    excel_file = kwargs.get('file')
    tab = kwargs.get('tab')
    pos = kwargs.get('pos', 0)
    book = load_workbook(excel_file)
    book._sheets.insert(pos, book._sheets.pop(book._sheets.index(book[tab])))
    book.save(excel_file)

def rename_excel_sheet(**kwargs):
    file = kwargs.get('file')
    sheet_map = kwargs.get('sheet_map', {
        'vs_profile': 'vs.profiles',
        'member': 'pool.members',
        'profile_list': 'profile.data' 
        })
    ss = load_workbook(file)
    for old, new in sheet_map.items():
        if new in ss.sheetnames:
            continue
        ss_sheet = ss[old]
        ss_sheet.title = new
    ss.save(file)

def run_task(task_info, daemon=False):  # run_task((func, (a1, a2), {'k1': v1, 'k2': v2}))
    task_args, task_kwargs = (), {}
    if callable(task_info):
        task_func = task_info
    else:
        task_func = task_info[0]
        if len(task_info) > 1:
            task_args = task_info[1]
        if len(task_info) > 2:
            task_kwargs = task_info[2]
    if task_args:
        thread_instance = threading.Thread(target=task_func, args=task_args, kwargs=task_kwargs, daemon=daemon)
    else:
        thread_instance = threading.Thread(target=task_func, kwargs=task_kwargs, daemon=daemon)
    thread_instance.start()
    return thread_instance

def parallel_func(**kwargs):
    func_list = kwargs.get('func')
    args_list = kwargs.get('args')
    kwargs_list = kwargs.get('kwargs')
    max_worker = kwargs.pop('max_worker', 10)

    if args_list and type(args_list) is not list:
        assert type(args_list) is tuple
    if kwargs_list and type(kwargs_list) is not list:
        assert type(kwargs_list) is dict
    total = max(len(func_list) if type(func_list) is list else 0, len(args_list) if type(args_list) is list else 0, len(kwargs_list) if type(kwargs_list) is list else 0)

    if type(func_list) is not list:
        func_list = [func_list] * total
    
    if args_list:
        if type(args_list) is not list:
            args_list = [args_list] * total
    else:
        args_list = [()] * total
    if kwargs_list:
        if type(kwargs_list) is not list:
            kwargs_list = [kwargs_list] * total
    else:
        kwargs_list = [{}] * total
    
    with ThreadPoolExecutor() as executor:        
        q, results = [], []
        run = 0
        for func, args, kwargs in zip(func_list, args_list, kwargs_list):
            q.append(executor.submit(func, *args, **kwargs))
            run += 1
            if run == max_worker:
                completed_q = []
                while True:
                    for p in q:
                        if p.done():
                            completed_q.append(p)
                            results.append(p.result())
                            run -= 1
                    if run < max_worker:
                        break
                    else:
                        sleep(1)
                [q.remove(p) for p in completed_q]
        if q:
            results += [p.result() for p in q]
    return results

# {'networkPolicyObject': ['obj1', 'obj2']}
# (csm_GetDeviceListByCapability, data_dir + fr'\csm.devices.xml'),
# policy name / dev map
# NAT policy name map
# if policy_type in policy_list:
# ISE-TACACS
"""                    
def hitcnt_pipeline(dev, **kwargs):
    parse_dev_info(dev, **kwargs)

        run_thread_parallel(
        hitcnt_pipeline,
        *fw_list, 
        max_workers=len(fw_list),
        dir=data_dir, 
        dev_os='cisco_asa',
        cmd='show access-list'
        )

    run_thread_parallel(
        get_single_f5_data,
        *f5_working_list, 
        dir=data_dir, 
        user=svc_usr, 
        pwd=svc_pwd,
        max_workers=16
        )

    refresh_token = ''
    if 'refreshToken' in token_response.keys():
        refresh_token = token_response['refresh_token']['token']

                                if _attr not in ['policies']:
                                    my_logger.info(f'{dev} {obj} {_attr} {fullPath} null')
                            if _attr in ['policies']:
                                my_logger.info(f'{dev} {obj} {_attr} {fullPath}')
                            my_logger.info(f'{dev} {obj} {re.search(r'/([^/]+)\?', req_url).group(1)} null')

for L5_item in L5_list:
    assert type(L5_item) is dict
    for L5_item_k, L5_item_v in L5_item.items():
        if L5_item_k in ['gid', 'name']:
            if debug:
                my_logger.info(L5_item_k, L5_item_v)

"""
                            # my_logger.info('\n')
                    # my_logger.info('\n')
                # my_logger.info('\n')

# refGIDs: 2142 lines, exceed excel cell limit of 32767
"""
    debug = kwargs.get('debug', False)
                if debug:
                    print(L2_item_k, L2_item_v)
                if debug:
                    print(L2_item_k, L2_item_v)
                            if debug:
                                print(L3_item_k, L3_item_v)
                            if debug:
                                print(L3_item_k, L3_item_v)

                                        if debug:
                                            print(L4_item_k, L4_item_v)
                                        if debug:
                                            print(L4_item_k, L4_item_v)
    assert type(L2_list) is list
        assert type(L2_item) is dict
                    assert type(L3_item) is dict
                                assert type(L3_item_v) is dict
                                assert type(L4_item) is dict
                                        assert type(L5_list) is list
                                                                
                                if debug:
                                    print('\n')
        if debug:
            print('\n')
"""    
# InterfaceNAT64ManualFirewallPolicy -> interfaceNATManualFirewallPolicy
# InterfaceNAT64ObjectFirewallPolicy -> interfaceNAT64ObjectFirewallPolicy
exit_flag = True
def csm_keep_alive(**kwargs):
    if not csm_ping(**kwargs):
        csm_Login(**kwargs)
    t_now = datetime.now().replace(microsecond=0)
    print(f"{str(t_now)}")
    next_t_1m = t_now.replace(second=0) + timedelta(minutes=1)
    sleep_sec = (next_t_1m - t_now).seconds
    sleep(sleep_sec)
    while True:
        if exit_flag:
            csm_Logout()
            break
        
        sleep(600)
        t_now = datetime.now().replace(microsecond=0)
        my_logger.info(f"{str(t_now)}")
        if not csm_ping(**kwargs):
            break

def csm_GetServiceInfo():
    url_func = 'GetServiceInfo'
    xml_func = 'getServiceInfoRequest'
    return csm_post(url_func=url_func, xml_func=xml_func)

def csm_GetDeviceListByCapability(**kwargs):
    retry = kwargs.get('retry', 3)
    url_func = 'getDeviceListByType'
    xml_func = 'deviceListByCapabilityRequest'
    dev_cap = kwargs.get('dev_cap', '*')  # firewall, ids, router, switch
    xml_args = f'<deviceCapability>{dev_cap}</deviceCapability>'
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args, retry=retry)

def csm_GetDeviceListByGroup(**kwargs):
    url_func = 'getDeviceListByGroup'
    xml_func = 'deviceListByGroupRequest'
    grp_path = kwargs.get('grp_path')
    item_str = ''
    for item in grp_path.split('/'):
        item_str += f'<pathItem>{item}</pathItem>'
    xml_args = f"<deviceGroupPath>{item_str}</deviceGroupPath>"
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args)

def csm_GetDeviceConfigByGID(**kwargs):
    gid = kwargs.get('gid')
    url_func = 'getDeviceConfigByGID'
    xml_func = 'deviceConfigByGIDRequest'
    xml_args = f'<gid>{gid}</gid>'
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args)

def csm_GetDeviceConfigByName(**kwargs):
    dev = kwargs.get('dev')
    url_func = 'getDeviceConfigByName'
    xml_func = 'deviceConfigByNameRequest'
    xml_args = f'<name>{dev}</name>'
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args)

def csm_GetPolicyConfigByName(**kwargs):
    url_func = 'getPolicyConfigByName'
    xml_func = 'policyConfigByNameRequest'
    policy_name = kwargs.get('policy_name')
    policy_type = kwargs.get('policy_type', 'DeviceAccessRuleFirewallPolicy')
    xml_args = f'<name>{policy_name}</name><policyType>{policy_type}</policyType>'
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args)

def csm_GetHitcountDetailsByGID(**kwargs):
    url_func = 'getHitcountDetailsByGID'
    xml_func = 'hitCountRequest'
    dev_gid = kwargs.get('dev_gid')
    policy_type = kwargs.get('policy_type', 'DeviceAccessRuleUnifiedFirewallPolicy')
    policy_rule_gid_list = kwargs.get('policy_rule_gid_list')
    policy_rules_str = ''
    for policy_rule_gid in policy_rule_gid_list[:max_policy_rules]:
        policy_rules_str += f"<policyRuleGID>{policy_rule_gid}</policyRuleGID>"
    xml_args = f"""
    <deviceGID>{dev_gid}</deviceGID>
    <policyType>{policy_type}</policyType>
    """ + policy_rules_str
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args)

def get_dev_policy_hit_xml(**kwargs):
    data_dir = kwargs.get('dir')
    policy_type = kwargs.get('policy_type', 'DeviceAccessRuleUnifiedFirewallPolicy')
    gid_dev_items=csm_get_gid_dev_items(dir=data_dir)
    for gid, dev in gid_dev_items:
        dev_policy_file = data_dir + fr"\{dev}.{policy_type}.xlsx"
        if not os.path.exists(dev_policy_file):
            my_logger.info(f"No {dev_policy_file}")
            continue
        policy_rule_gid_list = pd.read_excel(dev_policy_file, sheet_name='policy')['gid'].tolist()
        policy_rule_gid_list = sorted(list(set(policy_rule_gid_list)))
        policy_rule_gid_len = len(policy_rule_gid_list)
        batches = ceil(policy_rule_gid_len/max_policy_rules)
        for i in range(batches):
            batch_policy_rule_gid_list = policy_rule_gid_list[max_policy_rules * i: max_policy_rules * (i + 1)]
            result = csm_GetHitcountDetailsByGID(dev_gid=gid, policy_type=policy_type, policy_rule_gid_list=batch_policy_rule_gid_list)
            xml_file = data_dir + fr"\{dev}.{policy_type}.batch{i + 1}.xml"
            str2file(file=xml_file, str=result)

def csm_ExecDeviceReadOnlyCLICmds(**kwargs):
    dev = kwargs.get('dev')
    sh_cmd = kwargs.get('sh_cmd', 'version')
    svc = 'utilservice'
    url_func = 'execDeviceReadOnlyCLICmds'
    xml_func = 'execDeviceReadOnlyCLICmdsRequest'
    xml_args = f"""<deviceReadOnlyCLICmd>
    <deviceName>{dev}</deviceName>
    <cmd>show</cmd>
    <argument>{sh_cmd}</argument>
    </deviceReadOnlyCLICmd>"""
    return csm_post(svc=svc, url_func=url_func, xml_func=xml_func, xml_args=xml_args)


def csm_GetPolicyObjectByGID(**kwargs):
    url_func = 'getPolicyObjectByGID'
    xml_func = 'getPolicyObjectByGID'
    obj_gid_list = kwargs.get('obj_gid_list')
    xml_args = ''
    for obj_gid in obj_gid_list:
        xml_args += f'<gid>{obj_gid}</gid>'
    return csm_post(url_func=url_func, xml_func=xml_func, xml_args=xml_args)

def map_gid_in_policy(**kwargs):
    df = kwargs.get('df')
    sub_df = kwargs.get('sub_df')
    cols = ['gid', 'orderId','description','isEnabled','direction','permit','sectionName','policyName',
    'ruleNo','interfaceRoleObjectGIDs', 'sources', 'destinations', 'services']
    new_df = df[cols].fillna('')
    def map_gid(x):
        if not x:
            return []  # rule not enabled has empty obj
        x = eval(x)
        assert type(x) is dict
        assert len(x) in [1, 2]        
        gid_items = []
        info_items = []
        obj_type = ''
        if 'gid' in x.keys():
            obj_type = 'int'
            gid_items = x['gid']
            df = sub_df['InterfaceRolePolicyObject']
        else:
            for k, v in x.items():
                if k in ['ipData', 'serviceParameters']:
                    if not obj_type:                    
                        obj_type = 'net' if k == 'ipData' else 'svc'
                    if k == 'ipData':
                        if type(v) is str:
                            v = [v]
                    else:
                        if type(v) is dict:
                            v = [v]
                        for i in v:
                            assert i['sourcePort'] is None or i['sourcePort']['port'] == "1-65535"
                        v = [i['protocol'] + " " + i['destinationPort']['port'] for i in v]
                    info_items += v
                else:
                    assert k in ['networkObjectGIDs', 'serviceObjectGIDs']
                    if not obj_type:                    
                        obj_type = 'net' if 'net' in k else 'svc'
                    gid_items = v['gid']
                    if k in ['networkObjectGIDs']:
                        df = sub_df['NetworkPolicyObject']
                    else:
                        df = sub_df['ServicePolicyObject']
        
        if not gid_items:
            return info_items

        if type(gid_items) is not list:
            gid_items = [gid_items]
        for gid in gid_items:
            _df = df[df['gid'] == gid]
            if _df.empty:
                _df = df[df['parentGID'] == gid]
            if len(_df) != 1:
                assert len(_df) == 0
                my_logger.info(f"gid not found: {gid}")
                info_items.append(gid)
            else:
                name = _df['name'].iloc[0]
                if not name and obj_type == 'int':
                    name = _df['comment'].iloc[0]
                    if not name:
                        name = _df['pattern'].iloc[0]
                info_items.append(name)

        return info_items
    
    for item in ['interfaceRoleObjectGIDs', 'sources', 'destinations', 'services']:
        new_df[item] = new_df[item].apply(map_gid)
    return new_df

def policy_p0_render_gid_reference(**kwargs):  # L1 obj name, match fw config
    dev_regex = kwargs.get('dev_regex')
    data_dir = kwargs.get('dir')
    policy_type = kwargs.get('policy_type', 'DeviceAccessRuleUnifiedFirewallPolicy')
    
    print(f"\npolicy_p0_render_gid_reference\n")    
    df = pd.read_excel(data_dir + fr"\csm.groups.xlsx")[['gid', 'name']].drop_duplicates().sort_values('name', key=lambda col: col.str.lower())
    gid_dev_items = list(df.itertuples(index=False, name=None))
    for gid, dev in gid_dev_items:
        if re.search(dev_exclude_regex, dev) or (dev_regex and not re.search(dev_regex, dev)):
            # print(f"skip {dev}")
            continue
        dev_policy_file = data_dir + fr"\{dev}.{policy_type}.xlsx"
        new_dev_policy_file = dev_policy_file[:-4] + "p0.xlsx"
        if file_up_to_date(file=new_dev_policy_file):
            print(f"{Path(new_dev_policy_file).name} up to date")
            continue
        
        print(f"{Path(dev_policy_file).name}")
        xls = pd.ExcelFile(dev_policy_file)
        name_func_items = [
            ('InterfaceRolePolicyObject', None),
            ('PortListPolicyObject', None),
            ('ServicePolicyObject', None),
            ('NetworkPolicyObject', None),
            ('policy', map_gid_in_policy),
        ]
        new_df = {}
        for name, func in name_func_items:
            df = pd.read_excel(xls, name).fillna('')
            if func:
                new_df[name] = func(dir=data_dir, df=df, sub_df=new_df)
            else:
                new_df[name] = df

        with pd.ExcelWriter(new_dev_policy_file, mode='w') as writer:
            for name, _ in name_func_items:
                new_df[name].to_excel(writer, sheet_name=name, index=False)


def render_refGIDs(x):        
    result = ''
    if x:
        x_var = eval(x) if type(x) is str else x
        if type(x_var) is list:  # [{'gid':[gid1, gid2]}]
            if len(x_var) == 1:
                x_var = x_var[0]
                assert type(x_var) is dict
        if type(x_var) is dict:
            result = x_var['gid']            
    if not result: 
        result = x
    if result and type(result) is not list:
        result = [result] 
    return result


def load_item_from_ext_file(x, data_dir=None):
    assert data_dir
    if x and type(eval(x)) is dict:
        x_var = eval(x)
        if 'ext_file' in x_var.keys():
            assert len(x_var) == 1
            return eval(file2str(file=data_dir + fr"\{x_var['ext_file']}"))
    return x                

def render_nw(**kwargs):
    data_dir = kwargs.get('dir')
    df = kwargs.get('df')
    if 'fqdnData' in df.columns:
        fqdn_flag = True
        cols = ['gid', 'parentGID', 'name',	'comment', 'refGIDs', 'subType', 'ipData', 'fqdnData']
    else:
        fqdn_flag = False
        cols = ['gid',	'parentGID', 'name',	'comment', 'refGIDs', 'subType', 'ipData']
    new_df = df[cols].fillna('')
    new_df['info'] = ''
    new_df['refGIDs'] = new_df['refGIDs'].apply(load_item_from_ext_file, data_dir=data_dir)
    new_df['refGIDs'] = new_df['refGIDs'].apply(render_refGIDs)
    new_df['ipData'] = new_df['ipData'].apply(lambda x: x if not x else [x] if type(x) is str else eval(x))
    new_df.loc[new_df['ipData'] != '', 'info'] = new_df[new_df['ipData'] != ''].apply(lambda line: {line['name']: line['ipData']}, axis=1)
    
    if fqdn_flag:
        new_df['fqdnData'] = new_df['fqdnData'].apply(lambda x: x if not x else eval(x)['value'])
        new_df.loc[new_df['fqdnData'] != '', 'info'] = new_df[new_df['fqdnData'] != ''].apply(lambda line: {line['name']: line['fqdnData']}, axis=1)

    ref_df = new_df[new_df['refGIDs'] != '']
    # 00000000-0000-0000-0001-005022349850| L1 refGID 00000000-0000-0000-0004-458176072473
    for item in ref_df.itertuples():  # line of top level
        info = {}
        print(f"level 0 {item.gid, item.name}")
        for gid in item.refGIDs:  # L1 refGIDs  
            level = 1
            _df = new_df[new_df['gid'] == gid]  # line of gid in refGIDs
            if _df.empty:
                _df = new_df[new_df['parentGID'] == gid]
                print(f"{gid} in parentGID")
            assert len(_df) == 1
            name = _df['name'].iloc[0]            
            print(f"level {level} {gid, name}")

            sub_gid_table = []
            while True:
                sub_refGIDs = _df['refGIDs'].iloc[0]
                if sub_refGIDs:  # L2 refGIDs and further
                    level += 1
                    for sub_gid in sub_refGIDs:
                        sub_gid_table.append((level, sub_gid))
                        print(f"push level {level} {sub_gid}")
                ipData = _df['ipData'].iloc[0]
                if ipData:
                    info.setdefault(name, [])
                    info[name] += ipData
                else:
                    if fqdn_flag:
                        fqdnData = _df['fqdnData'].iloc[0]
                        if fqdnData:
                            info.setdefault(name, [])                            
                            info[name] += fqdnData
                
                if sub_gid_table:
                    level, sub_gid = sub_gid_table.pop(0)
                    _df = new_df[new_df['gid'] == sub_gid]
                    if _df.empty:
                        _df = new_df[new_df['parentGID'] == sub_gid]
                        print(f"{sub_gid} in parentGID")
                    assert len(_df) == 1
                    name = _df['name'].iloc[0]
                    print(f"pop level {level} {sub_gid, name}")
                else:
                    break
        new_df.loc[item.Index, 'info'] = [info]                
    return new_df

def render_int(**kwargs):
    df = kwargs.get('df')
    cols = ['gid', 'parentGID', 'name',	'pattern', 'comment']
    new_df = df[cols].fillna('')    
    
    def _map_int(line):
        name_items = ['name', 'pattern', 'comment']
        for name in name_items:
            if line[name] != '':
                return line[name] 
        return ''

    new_df['info'] = new_df.apply(_map_int, axis=1)
    return new_df

def render_po(**kwargs):
    df = kwargs.get('df')
    cols = ['gid',	'name',	'comment', 'port']
    new_df = df[cols].fillna('')
    new_df['port'] = new_df['port'].apply(lambda x: (int(eval(x)['startPort']['portNum']), int(eval(x)['endPort']['portNum'])))
    new_df['info'] = new_df.apply(lambda line: {line['name']: line['port']}, axis=1)
    return new_df

def render_svc_para(x):  # x is list
    result = []
    if x:
        for i in x:
            if 'destinationPort' in i.keys() and 'port' in i['destinationPort'].keys():
                svc = f"{i['protocol']}:{i['destinationPort']['port']}"
            else:
                svc = f"{i['protocol']}"
            result.append(svc)
    return result

def render_svc(**kwargs):
    df = kwargs.get('df')
    cols = ['gid',	'name',	'comment', 'refGIDs', 'subType', 'serviceParameters']
    new_df = df[cols].fillna('')
    new_df['refGIDs'] = new_df['refGIDs'].apply(lambda x: eval(x)['gid'] if x and type(eval(x)) is dict else x)
    new_df['refGIDs'] = new_df['refGIDs'].apply(lambda x: [x] if x and type(x) is not list else x)
    new_df['serviceParameters'] = new_df['serviceParameters'].apply(lambda x: x if not x else eval(x) if type(eval(x)) is list else [eval(x)])
    new_df['serviceParameters'] = new_df['serviceParameters'].apply(render_svc_para)
    new_df.loc[new_df['serviceParameters'] != '', 'info'] = new_df[new_df['serviceParameters'] != ''].apply(lambda line: {line['name']: line['serviceParameters']}, axis=1)

    ref_df = new_df[new_df['refGIDs'] != '']
    for item in ref_df.itertuples():
        info = {}
        for gid in item.refGIDs:
            _df = new_df[new_df['gid'] == gid]
            assert len(_df) == 1
            name = _df['name'].iloc[0]
            serviceParameters = _df['serviceParameters'].iloc[0]
            if serviceParameters:
                info[name] = serviceParameters
        new_df.loc[item.Index, 'info'] = [info]                
    return new_df

def render_policy(**kwargs):
    df = kwargs.get('df')
    sub_df = kwargs.get('sub_df')
    cols = ['gid', 'orderId','description','isEnabled','direction','permit','sectionName','policyName',
    'ruleNo','interfaceRoleObjectGIDs', 'sources', 'destinations', 'services']
    new_df = df[cols].fillna('')

    # {'serviceParameters': [{'protocol': 'tcp', 'sourcePort': None, 'destinationPort': {'port': '445'}}, 
    # {'protocol': 'tcp', 'sourcePort': None, 'destinationPort': {'port': '135'}}]}

    # {'serviceObjectGIDs': {'gid': '00000000-0000-0000-0000-502511174239'}, 
    # 'serviceParameters': {'protocol': 'tcp', 'sourcePort': None, 'destinationPort': {'port': '18891'}}}    
    def map_gid(x):
        x = eval(x)
        assert type(x) is dict and len(x) <= 2
        # non_gid_dict = None
        gid_items = []
        if 'gid' in x.keys():
            df = sub_df['InterfaceRolePolicyObject']
            gid_items = x['gid']
        else:
            for k, v in x.items():
                if k in ['ipData', 'serviceParameters']:
                    if k in ['ipData']:
                        if type(v) is str:
                            v = [v]                            
                        # assert type(v) is list                            
                        return v
                    else:
                        if type(v) is dict:
                            v = [v]
                        assert type(v) is list                            
                        return render_svc_para(v)
                    # non_gid_dict = {k:v}
                else:
                    assert k in ['networkObjectGIDs', 'serviceObjectGIDs']
                    gid_items = v['gid']
                    if k in ['networkObjectGIDs']:
                        df = sub_df['NetworkPolicyObject']
                    else:
                        df = sub_df['ServicePolicyObject']
        
        assert len(gid_items)
        if type(gid_items) is not list:
            gid_items = [gid_items]
        info_items = []
        for gid in gid_items:
            _df = df[df['gid'] == gid]
            if _df.empty:
                _df = df[df['parentGID'] == gid]
            if len(_df) != 1:
                print(gid)
                info_items.append(gid)
            else:
                info_items.append(_df['info'].iloc[0])
        # if non_gid_dict:
        #     info_items.append(non_gid_dict)
        return info_items
    
    for item in ['interfaceRoleObjectGIDs', 'sources', 'destinations', 'services']:
        new_df[item] = new_df[item].apply(map_gid)
    return new_df

def extract_ip_port(**kwargs):
    df = kwargs.get('df')
    cols = ['gid', 'orderId','description','isEnabled','direction','permit','sectionName','policyName',
    'ruleNo','interfaceRoleObjectGIDs', 'sources', 'destinations', 'services']
    new_df = df[cols].fillna('')

    def extract_info_v2(x):
        def _from_dict(v_dict):
            _result = []
            for k, v in v_dict.items():
                if type(v) is str:
                    assert v[0] not in ['[', '{']
                    _result.append(v)
                else:
                    assert type(v) is list
                    _result += v
            return _result

        def _from_list(v_list):
            _result = []
            for _item in v_list:
                if type(_item) is str:
                    if _item == '':
                        pass
                    elif _item[0] == '{':
                        _result += _from_dict(eval(_item))
                    else:
                        _result.append(_item)
                elif type(_item) is list:  # sources: [[{}], [{}]]
                    _result += _from_list(_item)
                else:
                    assert type(_item) is dict
                    _result += _from_dict(_item)
            return _result

        result = []
        if type(x) is dict:
            result += _from_dict(x)
        elif type(x) is list:
            result += _from_list(x)
        else:
            assert type(x) is str
            if x[0] == '{':
                result += _from_dict(eval(x))
            else:
                assert x[0] == '['
                result += _from_list(eval(x))
        return result
    
    for item in ['sources', 'destinations']:
        # new_df[item] = new_df[item].apply(extract_ip)
        new_df[item] = new_df[item].apply(extract_info_v2)  # test
    for item in ['services']:
        # new_df[item] = new_df[item].apply(extract_port)
        new_df[item] = new_df[item].apply(extract_info_v2)  # test
    return new_df

def policy_p1_render_gid_reference(**kwargs):  # leaf obj and content
    dev_regex = kwargs.get('dev_regex')
    data_dir = kwargs.get('dir')
    policy_type = kwargs.get('policy_type', 'DeviceAccessRuleUnifiedFirewallPolicy')

    print(fr"\policy_p1_render_gid_reference\n")

    df = pd.read_excel(data_dir + fr"\csm.groups.xlsx")[['gid', 'name']].drop_duplicates().sort_values('name', key=lambda col: col.str.lower())
    gid_dev_items = list(df.itertuples(index=False, name=None))
    name_func_items = [
        ('InterfaceRolePolicyObject', render_int),
        ('PortListPolicyObject', render_po),
        ('ServicePolicyObject', render_svc),
        ('NetworkPolicyObject', render_nw),
        ('policy', render_policy),
    ]
    for gid, dev in gid_dev_items:
        if (dev_exclude_regex and re.search(dev_exclude_regex, dev)) or (dev_regex and not re.search(dev_regex, dev)):
            # my_logger.info(f"skip {dev}")
            continue
        
        dev_policy_file = data_dir + fr"\{dev}.{policy_type}.xlsx"
        if not os.path.exists(dev_policy_file):
            my_logger.info(f"No {Path(dev_policy_file).name}")
            continue        
        
        print(f"{dev}")
        new_dev_policy_file = data_dir + fr"\{dev}.{policy_type}.p1.xlsx"
        if file_up_to_date(file=new_dev_policy_file):
            print(f"{Path(new_dev_policy_file).name} up to date")
            continue

        xls = pd.ExcelFile(dev_policy_file)
        new_df = {}
        for name, func in name_func_items:
            df = pd.read_excel(xls, name).fillna('')
            new_df[name] = func(dir=data_dir, df=df, sub_df=new_df)

        with pd.ExcelWriter(new_dev_policy_file, mode='w') as writer:
            for name, _ in name_func_items:
                for item in ['refGIDs', 'sources',	'destinations', 'info']:
                    save_item_in_ext_file(**kwargs, dev=dev, df=new_df[name], item=item)
                new_df[name].to_excel(writer, sheet_name=name, index=False)


def policy_p2_render_ip_port(**kwargs):  # content only for traffic match
    print("policy_p2_render_ip_port\n")
    dev_regex = kwargs.get('dev_regex')
    data_dir = kwargs.get('dir')
    policy_type = kwargs.get('policy_type', 'DeviceAccessRuleUnifiedFirewallPolicy')
    df = pd.read_excel(data_dir + fr"\csm.groups.xlsx")[['gid', 'name']].drop_duplicates().sort_values('name', key=lambda col: col.str.lower())
    gid_dev_items = list(df.itertuples(index=False, name=None))
    for gid, dev in gid_dev_items:
        if re.search(dev_exclude_regex, dev) or (dev_regex and not re.search(dev_regex, dev)):
            # my_logger.info(f"skip {dev}")
            continue
        
        dev_policy_file = data_dir + fr"\{dev}.{policy_type}.p1.xlsx"
        if not os.path.exists(dev_policy_file):
            my_logger.info(f"No {Path(dev_policy_file).name}")
            continue        
        
        print(f"{dev}")
        new_dev_policy_file = data_dir + fr"\{dev}.{policy_type}.p2.xlsx"
        if file_up_to_date(file=new_dev_policy_file):
            print(f"{Path(new_dev_policy_file).name} up to date")
            continue

        xls = pd.ExcelFile(dev_policy_file)
        name_func_items = [
            ('policy', extract_ip_port),
        ]
        new_df = {}
        for name, func in name_func_items:
            df = pd.read_excel(xls, name).fillna('')
            new_df[name] = func(df=df, sub_df=new_df)

        with pd.ExcelWriter(new_dev_policy_file, mode='w') as writer:
            for name, _ in name_func_items:
                # save_item_in_ext_file(df=new_df[name], item='refGIDs')
                # save_item_in_ext_file(df=new_df[name], item='info')
                new_df[name].to_excel(writer, sheet_name=name, index=False)

def csm_post_process(**kwargs):
    process_stages = [
        policy_p0_render_gid_reference, 
        policy_p1_render_gid_reference, 
        policy_p2_render_ip_port,
        ]
    stages = kwargs.get('stages', len(process_stages))
    for process_func in process_stages[:stages]:
        process_func(**kwargs)

def process_gtm_info_p1(**kwargs):
    data_dir = kwargs.get('dir', '.')
    dev = kwargs.get('dev')

    excel_file = data_dir + fr'\{dev}.xlsx'
    if file_up_to_date(file=excel_file) and check_tab_in_excel(file=excel_file, tab='p1'):
        print(f"{Path(excel_file).name} up to date and has p1 tab")
        # if get_tab_pos(file=excel_file, tab='p1') != 0:
        #     move_tab(file=excel_file, tab='p1')
        return
    
    top_cols = ['name']
    top = pd.read_excel(excel_file, sheet_name='top').fillna('')[top_cols]
    top['dc'] = top['name'].apply(lambda x: x.split('datacenter')[1].strip())
    top['subnet'] = top['name'].apply(lambda x: x.split(' server: ')[0].split(' ')[-1].strip())
    top_cols = ['dc', 'subnet']
    top = top[top_cols]

    dns_cols = ['name', 'partition', 'poolLbMode', 'pools']
    dns = pd.read_excel(excel_file, sheet_name='dns.a').fillna('')[dns_cols]
    dns.rename(columns={'name': 'dns', 'partition': 'dns.partition'}, inplace=True)
    dns['pools'] = dns['pools'].apply(eval)
    dns = dns.explode('pools')
    dns['pool'] = dns['pools'].apply(lambda x: '/' + x['partition'] + '/' + x['name'])
    dns.drop(['pools'], axis=1, inplace=True)
    
    pool_cols = ['fullPath', 'loadBalancingMode', 'alternateMode', 'fallbackMode']  # normal, static, last resort
    pool = pd.read_excel(excel_file, sheet_name='pool.a').fillna('')[pool_cols]
    pool.rename(columns={'fullPath': 'pool'}, inplace=True)
    
    pool_member_cols = ['fullPath',	'pool']
    pool_member = pd.read_excel(excel_file, sheet_name='pool.a.members').fillna('')[pool_member_cols]
    pool_member['server'] = pool_member['fullPath'].apply(lambda x: x.split(':')[0])
    pool_member['vs'] = pool_member['fullPath'].apply(lambda x:  x.split(':')[1] if ':' in x else '')
    pool_member_cols = ['pool', 'server', 'vs']
    pool_member = pool_member[pool_member_cols]

    server_cols = ['fullPath', 'datacenter', 'product']
    server = pd.read_excel(excel_file, sheet_name='server').fillna('')[server_cols]
    server.rename(columns={'fullPath': 'server'}, inplace=True)

    vs_cols = ['name',	'destination', 'server']
    vs = pd.read_excel(excel_file, sheet_name='server.virtual-servers').fillna('')[vs_cols]
    vs.rename(columns={'name': 'vs'}, inplace=True)

    gtm = pd.merge(dns, pool, on='pool')
    tab_key_map = [
        (pool_member, 'pool'), 
        (vs, ['server', 'vs']), 
        (server, 'server')
        ]
    for a, k in tab_key_map:
        gtm = pd.merge(gtm, a, on=k)
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        gtm.to_excel(writer, sheet_name='p1', index=False)
    
    move_tab(file=excel_file, tab='p1')
    print(f"save {Path(excel_file).name} p1")


def get_pools_by_rule(**kwargs):
    rule = kwargs.get('rule')
    df = kwargs.get('df')
    rule_text = df.loc[df['fullPath'] == rule, 'apiAnonymous'].iloc[0]
    pools = re.findall(r'^ +pool (/.+)$', rule_text, flags=re.MULTILINE)
    return pools

def get_pools_by_rules(**kwargs):
    rules = kwargs.get('rules')
    df = kwargs.get('df')
    pools = []
    for rule in rules:
        pools += get_pools_by_rule(rule=rule, df=df)
    return pools

def get_members_by_pool(**kwargs):
    pool = kwargs.get('pool')
    df = kwargs.get('df')
    pool_in_member = pool.replace('/', '~')
    return df.loc[df['selfLink'].str.contains(fr"/{pool_in_member}/")]

def process_ltm_info_p1(**kwargs):
    data_dir = kwargs.get('dir', '.')
    dev = kwargs.get('dev')

    excel_file = data_dir + fr'\{dev}.xlsx'
    if file_up_to_date(file=excel_file) and check_tab_in_excel(file=excel_file, tab='p1'):
        print(f"{Path(excel_file).name} up to date and has p1 tab")
        if get_tab_pos(file=excel_file, tab='p1') != 0:
            move_tab(file=excel_file, tab='p1')
        return

    vs = pd.read_excel(excel_file, sheet_name='vs').fillna('')
    member = pd.read_excel(excel_file, sheet_name='pool.members').fillna('')
    rule = pd.read_excel(excel_file, sheet_name='rule').fillna('')    
    cols = ['vs', 'ip', 'protocol', 'port', 'pool', 'member', 'node ip', 'node port']    
    ltm = pd.DataFrame(columns=cols)
    i = 0
    for item in vs.itertuples():
        vs_fullPath = item.fullPath
        vs_ip_port = item.destination.split('/')[-1]
        if re.search(r'(\d+\.){3}\d+(%\d+)?:\d+', vs_ip_port):
            vs_ip, vs_port = vs_ip_port.split(':')
        else:
            vs_ip, vs_port = vs_ip_port.split('.')
        vs_protocol = item.ipProtocol
        if 'pool' in vs.columns:
            vs_pool = item.pool
        else:
            vs_pool = ''
        if vs_pool:
            vs_pools = [vs_pool]
        else:
            if 'rules' in vs.columns:
                rules = item.rules
            else:
                rules = ''
            if rules:
                vs_pools = get_pools_by_rules(rules=eval(rules), df=rule)
            else:
                vs_pools = []
                my_logger.info(f"{dev} {vs_fullPath} no pool or rules")
        for vs_pool in vs_pools:
            pool_member = get_members_by_pool(pool=vs_pool, df=member)
            if pool_member.empty:
                my_logger.info(f"{dev} {vs_pool} no member")
            else:
                for item in pool_member.itertuples():
                    pool_node, node_port = item.name.split(':')
                    node_ip = item.address
                    ltm.loc[i] = [vs_fullPath, vs_ip, vs_protocol, vs_port, vs_pool, pool_node, node_ip, node_port]
                    i += 1        
    
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        ltm.to_excel(writer, sheet_name='p1', index=False)
    print(f"save {Path(excel_file).name} p1")

def process_single_f5_data(dev, **kwargs):
    data_dir = kwargs.get('dir')
    f = data_dir + fr"\{dev}.xlsx"
    if os.path.exists(f):
        if '-ltm-' in dev:
            process_func = process_ltm_info_p1 
        else:
            process_func = process_gtm_info_p1 
        try:
            process_func(dev=dev, dir=data_dir)
        except Exception as e:
            my_logger.info(f'{dev}:\n{e}')
    else:
        my_logger.info(f"no {f}")
   
def get_single_f5_ip(*args, **kwargs):
    f = args[0]
    cols = ['f5', 'type', 'name', 'ip', 'protocol', 'port']
    process_info = [
        {
            'sheet': 'self',
            'type': 'self',
            'rename': {'name':'short_name', 'fullPath': 'name', 'address': 'ip'}
        },
        {
            'sheet': 'p1',
            'type': 'vs',
            'rename': {'vs': 'name'}
        }
        ]
    f5_ip_df = pd.DataFrame(columns=cols)
    if os.path.exists(f):
        f5 = Path(f).name.split('.')[0]
        if '-ltm-' in f5:
            try:
                for item in process_info:
                    df = pd.read_excel(f, sheet_name=item['sheet']).rename(columns=item['rename']).drop_duplicates()
                    df = df[[col for col in cols if col in df.columns]]
                    df['ip'] = df['ip'].apply(lambda x: x.split('/')[0].split('%')[0])
                    item_type = item['type']
                    if item_type in ['self']:
                        df = df.assign(f5=f5, type=item_type, protocol='', port='')
                    else:
                        df = df.assign(f5=f5, type=item_type)        
                    f5_ip_df = pd.concat([f5_ip_df, df])
            except Exception as e:
                my_logger.info(f'{f}:\n{e}')
    return f5_ip_df

def get_all_f5_ip(**kwargs):
    data_dir = kwargs.get('dir')
    devs = kwargs.get('devs')
    f = data_dir + fr'\mddr.f5.xlsx'
    if file_up_to_date(file=f):
        print(f"{Path(f).name} up to date")
        return

    f5_ip_df = pd.concat(
        run_thread_parallel(
            get_single_f5_ip, 
            *[data_dir + fr"\{dev}.xlsx" for dev in devs],
            max_workers=len(f5_working_list)
            ),
        ignore_index=True
        )

    if len(f5_ip_df):
        f5_ip_df.sort_values(by=['f5', 'type', 'name']).to_excel(f, index=False)

       
run_thread_parallel(
    process_single_f5_data, 
    *f5_working_list, 
    dir=data_dir, 
    max_workers=len(f5_working_list)
    )

get_all_f5_ip(
    dir=data_dir, 
    devs=f5_working_list,        
)
    
# python -m zipapp csm -m csm_api:csm_job -o dist\csm_api.pyz

cmds_template = Template("""
password:
$pwd
Press Return to begin session.

$fw>
en
Password:
$enable_pwd
$fw#
term pager 0
$fw#
show access-list
$fw#
"""
)

def get_dev_info_expect(dev, ip, **kwargs):
    usr = kwargs.get('usr')
    pwd = kwargs.get('pwd')
    enable_pwd = kwargs.get('enable_pwd')

    dev_info = ''
    try:
        cmds = cmds_template.substitute(fw=dev, pwd=pwd, enable_pwd=enable_pwd)
        expect_send_list = [line.strip() for line in cmds.strip().split('\n')]
        child = wexpect.spawn(f'{putty_exe} -P 22 {usr}@{ip}')
        cmd_state = 'expect'
        timeout = timeout_short
        for line in expect_send_list:
            if cmd_state == 'expect':
                # print(child.before)
                child.expect(line, timeout=timeout)
                # print(child.after)
                cmd_state = 'send'
            else:
                child.sendline(line)
                cmd_state = 'expect'
                timeout = timeout_long if line in ['show access-list'] else timeout_short

        dev_info = child.before
        child.sendline('exit')
        child.wait()
    except Exception as e:
        my_logger.info(f'\n{dev}:\n{e}')
        dev_info = ''
    return dev_info

# dev_info = get_dev_info_expect(fw, ip, usr=usr, pwd=rsa_pin + getpass('RSA Token(6):'), enable_pwd=enable_pwd)
import ipaddress

def explode_csm_fw_rule(dev, **kwargs):
    func = 'explode_csm_fw_rule'
    data_dir = kwargs.get('dir')
    for suffix in ['hit.vs.csm', 'csm.p0.explode']:
        out_fname = data_dir + fr'\{dev}.{suffix}.xlsx'
        if file_up_to_date(file=out_fname):
            print(f"{Path(out_fname).name} up to date")
            return

    policy_p0_render_gid_reference(dir=data_dir, dev_regex=dev)
    fname = data_dir + fr"\{dev}.DeviceAccessRuleUnifiedFirewallPolicy.p0.xlsx"
    df = pd.read_excel(fname, "policy")
    os.remove(fname)
    df = df[df['isEnabled'] == True]
    cols = ['sources', 'destinations', 'services']
    for col in cols:
        df[col] = df[col].apply(lambda x: x if type(x) is list else eval(x))
        df = df.explode(col)
    df.to_excel(out_fname, index=False)
    print(f'{dev, func} completed')

def service_2_protocol_svc(row):
    service = row['services']
    protocol, port = '', ''
    if service in svc_protocol_port_map.keys():
        protocol, port = svc_protocol_port_map[service]
    else:
        if re.fullmatch(r'^(tcp|udp) \d+$', service):
            protocol, port = service.split(' ')
            k = protocol + ':' + port
            if k in protocol_port_svc_map.keys():
                port = protocol_port_svc_map[k]
        elif re.fullmatch(r'^(tcp|udp) (\d+)-(\d+)$', service):
            protocol, port_range = service.split(' ')
            port_start, port_end = port_range.split('-')
            k = protocol + ':' + port_start
            if k in protocol_port_svc_map.keys():
                port_start = protocol_port_svc_map[k]
            k = protocol + ':' + port_end
            if k in protocol_port_svc_map.keys():
                port_end = protocol_port_svc_map[k]
            port = port_start + '-' + port_end
        else:
            port = service
    return pd.Series([protocol, port])

def net_obj_2_fw_cfg(**kwargs):
    net_obj = kwargs.get('net_obj')
    full_flag = kwargs.get('full_flag', True)
    if re.fullmatch(r'(\d+.){3}\d+/\d+', net_obj):
        ip_net = ipaddress.IPv4Network(net_obj, strict=False)
        info = f"{ip_net.network_address} {ip_net.netmask}"
    elif re.fullmatch(r'(\d+.){3}\d+', net_obj):
        if full_flag:
            info = "host " + net_obj
        else:
            info = net_obj
    else:
        if full_flag:
            info = "object.* " + re.escape(net_obj)
        else:
            info = net_obj
    return info

def net_obj_2_cfg(net_obj, **kwargs):
    full_flag = kwargs.get('full_flag', True)
    if net_obj in net_obj_map.keys():
        cfg = net_obj_map[net_obj]
    else: 
        cfg = net_obj_2_fw_cfg(net_obj=net_obj, full_flag=full_flag)
    return cfg
    
def map_explode_csm_fw_rule_2_cfg(dev, **kwargs):
    func = 'map_explode_csm_fw_rule_2_cfg'
    data_dir = kwargs.get('dir')
    for suffix in ['hit.vs.csm', 'csm.p0.explode.2.cfg']:
        out_fname = data_dir + fr'\{dev}.{suffix}.xlsx'
        if file_up_to_date(file=out_fname):
            print(f"{Path(out_fname).name} up to date")
            return    

    fname = data_dir + fr"\{dev}.csm.p0.explode.xlsx"
    df = pd.read_excel(fname)
    os.remove(fname)
    df['sources'] = df['sources'].apply(net_obj_2_cfg, full_flag=False, axis=1)
    df['destinations'] = df['destinations'].apply(net_obj_2_cfg, full_flag=False, axis=1)
    df[['protocol', 'svc']] = df.apply(service_2_protocol_svc, axis=1)
    df = df.explode('protocol')
    df.to_excel(out_fname, index=False)
    print(f'{dev, func} completed')

def concat_svc(row):
    port_start = row['dst_port_range_start']
    if port_start:
        port_end = row['dst_port_range_end']
        result = f"{port_start}-{port_end}"
    else:
        result = ''.join(row)
    return result

def concat_src(row):    	
    nw = row['src_network']
    if nw:        
        result = f"{nw} {row['src_mask']}"
    else:
        result = ''.join(row)
    return result

def concat_dst(row):
    nw = row['dst_network']
    if nw:        
        result = f"{nw} {row['dst_mask']}"
    else:
        result = ''.join(row)
    return result

def concat_hit_file_src_dst_svc(dev, **kwargs):
    func = 'concat_hit_file_src_dst_svc'
    data_dir = kwargs.get('dir')
    for suffix in ['hit.vs.csm', 'acl.hit.2.rule']:
        out_fname = data_dir + fr'\{dev}.{suffix}.xlsx'
        if file_up_to_date(file=out_fname):
            print(f"{Path(out_fname).name} up to date")
            return

    fname = data_dir + fr"\{dev}.acl.hit.xlsx"
    df = pd.read_excel(fname, dtype=str)
    os.remove(fname)

    svc_dst_cols = ['dst_port', 'dst_port_range_start', 'dst_port_range_end', 'dst_port_grp', 'dst_port_object', 'dst_icmp_type']
    src_cols = [col for col in df.columns if 'src' in col]
    dst_cols = list(set([col for col in df.columns if 'dst' in col]) - set(svc_dst_cols))
    svc_cols = [col for col in df.columns if 'svc' in col] + svc_dst_cols
    df['sources'] = df[src_cols].fillna('').apply(concat_src, axis=1)
    df['destinations'] = df[dst_cols].fillna('').apply(concat_dst, axis=1)
    df['svc'] = df[svc_cols].fillna('').apply(concat_svc, axis=1)    
    df.to_excel(out_fname, index=False)
    print(f'{dev, func} completed')

def concat_explode_csm_fw_rule_cfg_n_hit(dev, **kwargs):
    func = 'concat_explode_csm_fw_rule_cfg_n_hit'
    data_dir = kwargs.get('dir')
    out_fname = data_dir + fr"\{dev}.hit.vs.csm.xlsx"
    if file_up_to_date(file=out_fname):
        my_logger.info(f"{Path(out_fname).name} up to date")
        return

    fname = data_dir + fr"\{dev}.csm.p0.explode.2.cfg.xlsx"
    csm_2_hit_df = pd.read_excel(fname)
    os.remove(fname)
    fname = data_dir + fr"\{dev}.acl.hit.2.rule.xlsx"
    hit_2_rule_df = pd.read_excel(fname)
    os.remove(fname)
    
    match_cols = ['sources', 'destinations', 'protocol', 'svc']
    cols = ['line_num'] + match_cols + ['hit_count', 'action']

    df = pd.merge(csm_2_hit_df[match_cols + ['gid']], hit_2_rule_df[cols], on=match_cols, how='right')
    df[['gid', 'line_num'] + match_cols + ['hit_count', 'action']].to_excel(out_fname, index=False)
    print(f'{dev, func} completed')   

from fw_hit import *

net_obj_map = {
    'All-Addresses': 'any',
    'All-IPv4-Addresses': 'any4',
}

svc_protocol_port_map = {
    'IP': ('ip', ''),
    'ICMP': ('icmp', ''),
    'PIM': ('pim', ''),
    'TCP': ('tcp', ''),
    'UDP': ('udp', ''),
    'Bootpc': ('udp', 'bootpc'),
    'Bootps': ('udp', 'bootps'),
    'Citrix-ICA': ('tcp', 'citrix-ica'),
    'DHCP-Relay': ('udp', 'bootps'),
    'DNS-UDP': ('udp', 'domain'),
    'DNS-TCP': ('tcp', 'domain'),
    'FTP': ('tcp', 'ftp'),
    'FTP-Data': ('tcp', 'ftp-data'),
    'HTTP': ('tcp', 'www'),
    'HTTPS': ('tcp', 'https'),
    'IMAP4': ('tcp', 'imap4'),
    'LDAP': ('tcp', 'ldap'),
    'LDAPS': ('tcp', 'ldaps'),
    'MS-SQL-Server': ('tcp', '1433'),
    'MS-SQL-Monitor': (['tcp', 'udp'], '1434'),
    'Microsoft-ds': (['tcp', 'udp'], '445'),
    'Nbsession': ('tcp', 'netbios-ssn'),
    'Nbdatagram': ('udp', 'netbios-ns'),  # CSM definition TCP 137 is name service, 
    'Nbname': ('udp', 'netbios-dgm'),  # CSM definition TCP 138 is datagram service, 
    'NFS-TCP': ('tcp', 'nfs'),
    'NTP-TCP': ('tcp', '123'),
    'NTP-UDP': ('udp', 'ntp'),
    'SIP': (['tcp', 'udp'], 'sip'),
    'SMTP': ('tcp', 'smtp'),
    'SNMP': ('udp', 'snmp'),
    'SNMP-Trap': ('udp', 'snmptrap'),
    'SSH': ('tcp', 'ssh'),
    'Sun-RPC-TCP': ('tcp', 'sunrpc'),
    'Syslog': ('udp', 'syslog'),
    'Telnet': ('tcp', 'telnet'),
    'TFTP-TCP': ('tcp', '69'),
}

protocol_port_svc_map = {  # port range
    'udp:67': 'bootps',
    'udp:69': 'tftp',
    'tcp:80': 'www',
    'tcp:443': 'https',
    'tcp:1494': 'citrix-ica',
    'tcp:1521': 'sqlnet',
}

def hitcnt_pipeline(dev, **kwargs):
    parse_dev_info(dev, **kwargs)
    concat_hit_file_src_dst_svc(dev, **kwargs)
    explode_csm_fw_rule(dev, **kwargs)
    map_explode_csm_fw_rule_2_cfg(dev, **kwargs)
    concat_explode_csm_fw_rule_cfg_n_hit(dev, **kwargs)
